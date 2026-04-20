import openpyxl
from openpyxl import Workbook
from copy import copy
from pathlib import Path

def row_split(output_path):
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Leave the original processed workbook alone when it already fits within
    # the 500-row POS limit (header included).
    if ws.max_row <= 500:
        return
        
    file_count = 1
    dest_row = 2
    row_count = 0
    new_wb = Workbook()
    new_ws = new_wb.active
    input_path = Path(output_path)
    
    # Copy the header row once so every split workbook starts with the same columns.
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    new_ws.append(headers)
        
    # Copy data rows from the styled source workbook. row_count tracks only data rows,
    # so 499 here means each split file stays within the 500-row Excel limit once the
    # header row is included.
    for row in ws.iter_rows(min_row=2):
        if row_count >= 499:
            save_path = input_path.with_name(input_path.stem + "_" + str(file_count) + input_path.suffix)
            new_wb.save(save_path)
            # Start a fresh workbook chunk and reset the destination row tracking.
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.append(headers) # Re-add headers
            dest_row = 2
            file_count += 1
            row_count = 0

        for col_num, source_cell in enumerate(row, start=1):
            # Source rows from the original workbook map onto new local row numbers in
            # each split workbook, so copy cell-by-cell into the current destination row.
            target_cell = new_ws.cell(row=dest_row, column=col_num)
            target_cell.value = source_cell.value
            target_cell.fill = copy(source_cell.fill)

   
        dest_row += 1
        row_count += 1

        
    # Save the final chunk after the row loop finishes.
    save_path = input_path.with_name(input_path.stem + "_" + str(file_count) + input_path.suffix)
    new_wb.save(save_path)

    
