import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def convert_to_excel(df, file_name):
    # Keep the export in the same folder as the source file and append "_processed".
    input_path = Path(file_name)
    output_path = input_path.with_name(input_path.stem + "_processed" + input_path.suffix)

    # Write the processed DataFrame without the pandas index column.
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    return output_path


def apply_highlights(file_path, duplicates, prefilled):
    # Re-open the exported workbook so highlight styles can be applied to real Excel cells.
    wb = load_workbook(file_path)
    ws = wb.active

    yellow_fill = PatternFill(bgColor='FFFF00', fill_type="solid")
    blue_fill = PatternFill(bgColor='48CAE4', fill_type="solid")

    # Duplicate-review rows are highlighted across the full exported row.
    for idx in duplicates.index:
        excel_row = idx + 2
        for cell in ws[excel_row]:
            cell.fill = yellow_fill

    # Prefilled values highlight only the exact cell recorded in the tuple.
    for idx, column, value in prefilled:
        excel_row = idx + 2
        for header_cell in ws[1]:
            if header_cell.value == column:
                excel_col = header_cell.column

                target_cell = ws.cell(row=excel_row, column=excel_col)
                target_cell.fill = blue_fill
                break

    wb.save(file_path)

    return file_path


def export_menu_items(df, file_name, duplicates, prefilled):
    # Coordinate the export flow so main.py can call one function.
    output_path = convert_to_excel(df, file_name)
    apply_highlights(output_path, duplicates, prefilled)
    return output_path

